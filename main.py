from pydub import AudioSegment
from openpyxl import load_workbook
import csv

import pdb

CSV_PATH = 'output/labels.csv'

label_count = {
    'firearm': 0,
    'shot_85': 0,
    'shot_105': 0,
    'incoming_85': 0,
    'incoming_105': 0,
    'ags': 0,
    'unknown': 0
}

def extract_sounds(wav_file, ws):
    if not hasattr(extract_sounds, "labeled"): 
        extract_sounds.labeled = 0 
    if not hasattr(extract_sounds, "unknown"): 
        extract_sounds.unknown = 0 
    # pdb.set_trace()

    label_csv = open(CSV_PATH, 'a', newline='')
    csv_writer = csv.writer(label_csv)

    frame_val = ws['B3'].value
    event = ws['A3'].value
    row = 3

    while frame_val is not None or event is not None:
        event = event.strip().replace(" ", "_")
        label = get_label(event)

        frame_val = int(frame_val)
        start_time = (frame_val / frame_rate) * 1000
        end_time = start_time + 1500
        segment = wav_file[start_time:end_time]

        if (label == 'unknown'):
            extract_sounds.unknown += 1
            segment.export(f"output/unknown/{extract_sounds.unknown}_{label}.wav", format="wav")
        else:
            extract_sounds.labeled += 1
            segment.export(f"output/labeled/{extract_sounds.labeled}_{label}.wav", format="wav")
            csv_writer.writerow([label])
        
        row += 1
        event = ws['A' + str(row)].value
        frame_val = ws['B' + str(row)].value
    
    label_csv.close()


def get_label(event):
    label = ''
    if event.find('Стрілецька') != -1:
        label = 'firearm'
    elif event.find('Постріл') != -1:
        if event.find('85') != -1:
            label = 'shot_85'
        else:
            label = 'shot_105'
    elif event.find('Прихід') != -1:
        if event.find('85') != -1:
            label = 'incoming_85'
        else:
            label = 'incoming_105'
    elif event.find('АГС') != -1:
        label = 'ags'
    else:
        label = 'unknown'
    
    label_count[label] += 1
    return label


if __name__ == "__main__":
    open(CSV_PATH, 'w').close()
    
    wav_files = ["L_12_21.wav", "L_12_31.wav", "Vuho2.wav", "Vuho10_12_20.wav", "Vuho10_12_30.wav", "Vuho11.wav"]
    xlsx_sheets = ["laptop_12_21_22", "laptop_12_31_22", "V2", "V10_12_20_35", "V10_12_30_35", "V11"]
    # wav_files = ["Vuho2.wav"]
    # xlsx_sheets = ["V2"]

    for i in range(0, len(wav_files)):
        wav_file = AudioSegment.from_wav(f"audio-src/{wav_files[i]}")
        frame_rate = wav_file.frame_rate
        print("Frame rate ", frame_rate)
        wav_wb = load_workbook(filename=f"events.xlsx")
        wav_ws = wav_wb[xlsx_sheets[i]]

        extract_sounds(wav_file, wav_ws)
        print("Extracted for " + wav_files[i])
    
    print(label_count)
